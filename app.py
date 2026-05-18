import os
import io
import json
import secrets
from datetime import datetime, timedelta
from pathlib import Path
from functools import wraps
from concurrent.futures import ThreadPoolExecutor

from flask import (Flask, request, jsonify, render_template, send_file,
                   session, redirect, url_for, flash)
from dotenv import load_dotenv
from werkzeug.security import check_password_hash
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

import database
import extractor

load_dotenv(Path(__file__).parent / ".env", override=True)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY") or os.urandom(32)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024

# Flask-Mail (optional — graceful fallback if not installed/configured)
try:
    from flask_mail import Mail, Message as MailMessage
    app.config["MAIL_SERVER"]   = os.getenv("MAIL_SERVER", "smtp.gmail.com")
    app.config["MAIL_PORT"]     = int(os.getenv("MAIL_PORT", 587))
    app.config["MAIL_USE_TLS"]  = True
    app.config["MAIL_USERNAME"] = os.getenv("MAIL_USERNAME", "")
    app.config["MAIL_PASSWORD"] = os.getenv("MAIL_PASSWORD", "")
    app.config["MAIL_DEFAULT_SENDER"] = os.getenv("MAIL_FROM", os.getenv("MAIL_USERNAME", "noreply@example.com"))
    mail = Mail(app)
    MAIL_ENABLED = bool(os.getenv("MAIL_USERNAME"))
except ImportError:
    mail = None
    MAIL_ENABLED = False

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
executor = ThreadPoolExecutor(max_workers=3)

database.init_db()


# ── Auth helpers ──────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def superadmin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("role") != "superadmin":
            return "Forbidden — superadmin only.", 403
        return f(*args, **kwargs)
    return decorated


def current_user():
    return session.get("user_id"), session.get("role")


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() == "pdf"


# ── Public page routes ────────────────────────────────────────────────────────

@app.route("/")
def landing():
    return render_template("landing.html")


@app.route("/pricing")
def pricing():
    return render_template("pricing.html",
        razorpay_monthly  = database.get_site_setting("razorpay_url_monthly"),
        razorpay_annual   = database.get_site_setting("razorpay_url_annual"),
        razorpay_lifetime = database.get_site_setting("razorpay_url_lifetime"),
    )


@app.route("/reseller")
def reseller_landing():
    hubspot_embed_code = database.get_site_setting("hubspot_form_code")
    return render_template("reseller_landing.html", hubspot_embed_code=hubspot_embed_code)


@app.route("/contact")
def contact():
    content = database.get_site_setting("contact_us_content")
    return render_template("contact.html", content=content)


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = database.get_user_by_username(username)
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"]  = user["id"]
            session["username"] = user["username"]
            session["role"]     = user["role"]
            return redirect(url_for("dashboard"))
        error = "Invalid username or password."
    return render_template("login.html", error=error)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        user = database.get_user_by_email(email)
        if user:
            token = secrets.token_urlsafe(32)
            expires_at = (datetime.utcnow() + timedelta(hours=1)).strftime("%Y-%m-%d %H:%M:%S")
            database.create_reset_token(user["id"], token, expires_at)
            reset_url = url_for("reset_password", token=token, _external=True)

            if MAIL_ENABLED and mail:
                try:
                    msg = MailMessage(
                        subject="Password Reset — Resume Extractor",
                        recipients=[email],
                        body=f"Click the link below to reset your password (valid 1 hour):\n\n{reset_url}\n\nIf you did not request this, ignore this email.",
                    )
                    mail.send(msg)
                except Exception as e:
                    print(f"Mail send error: {e}")
                    # Dev fallback: flash the link
                    flash(f"(Dev) Reset link: {reset_url}", "info")
            else:
                # Dev fallback: flash the reset link directly
                flash(f"Dev mode — reset link: {reset_url}", "info")

        flash("If that email is registered, you'll receive a reset link.", "success")
        return redirect(url_for("forgot_password"))
    return render_template("forgot_password.html")


@app.route("/reset-password/<token>", methods=["GET", "POST"])
def reset_password(token):
    row = database.get_reset_token(token)
    if not row:
        flash("Invalid or expired reset link.", "danger")
        return redirect(url_for("login"))

    # Check expiry
    try:
        expires_at = datetime.strptime(row["expires_at"], "%Y-%m-%d %H:%M:%S")
        if datetime.utcnow() > expires_at:
            database.delete_reset_token(token)
            flash("Reset link has expired. Please request a new one.", "danger")
            return redirect(url_for("forgot_password"))
    except Exception:
        database.delete_reset_token(token)
        flash("Invalid reset link.", "danger")
        return redirect(url_for("login"))

    if request.method == "POST":
        new_password = request.form.get("new_password", "")
        confirm      = request.form.get("confirm_password", "")
        if not new_password:
            return render_template("reset_password.html", token=token, error="Password cannot be empty.")
        if new_password != confirm:
            return render_template("reset_password.html", token=token, error="Passwords do not match.")
        database.update_user_password(row["user_id"], new_password)
        database.delete_reset_token(token)
        flash("Password reset successfully. Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("reset_password.html", token=token)


# ── Page routes ───────────────────────────────────────────────────────────────

@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")


@app.route("/upload")
@login_required
def upload_page():
    return render_template("upload.html")


@app.route("/projects")
@login_required
def projects_page():
    return render_template("projects.html")


@app.route("/projects/<int:project_id>")
@login_required
def project_detail_page(project_id):
    user_id, role = current_user()
    project = database.get_project_by_id(project_id, user_id, role)
    if not project:
        return "Project not found or access denied.", 404
    return render_template("project_detail.html", project=project)


# ── Superadmin: user management ───────────────────────────────────────────────

@app.route("/admin/data")
@superadmin_required
def admin_data():
    return render_template("admin_data.html")


@app.route("/admin/site-settings", methods=["GET", "POST"])
@superadmin_required
def admin_site_settings():
    if request.method == "POST":
        for key in ("hubspot_form_code", "contact_us_content",
                    "razorpay_url_monthly", "razorpay_url_annual", "razorpay_url_lifetime"):
            database.set_site_setting(key, request.form.get(key, ""))
        flash("Site settings saved.", "success")
        return redirect(url_for("admin_site_settings"))
    return render_template("admin_site_settings.html",
        hubspot_form_code    = database.get_site_setting("hubspot_form_code"),
        contact_us_content   = database.get_site_setting("contact_us_content"),
        razorpay_url_monthly  = database.get_site_setting("razorpay_url_monthly"),
        razorpay_url_annual   = database.get_site_setting("razorpay_url_annual"),
        razorpay_url_lifetime = database.get_site_setting("razorpay_url_lifetime"),
    )


@app.route("/admin/leads")
@superadmin_required
def admin_leads():
    leads = database.get_all_leads()
    return render_template("admin_leads.html", leads=leads)


@app.route("/api/admin/users-by-role")
@superadmin_required
def api_admin_users_by_role():
    role = request.args.get("role", "").strip()
    if role not in ("superadmin", "reseller", "customer"):
        return jsonify([])
    return jsonify(database.get_users_by_role(role))


@app.route("/api/admin/projects")
@superadmin_required
def api_admin_projects():
    owner_role = request.args.get("role", "").strip()
    projects = database.get_admin_projects(owner_role if owner_role else None)
    return jsonify(projects)


@app.route("/admin/users")
@superadmin_required
def admin_users():
    users = database.get_all_users()
    resellers = [u for u in users if u["role"] == "reseller"]
    return render_template("admin_users.html", users=users, resellers=resellers)


@app.route("/admin/users/create", methods=["POST"])
@superadmin_required
def admin_create_user():
    username        = request.form.get("username", "").strip()
    email           = request.form.get("email", "").strip()
    password        = request.form.get("password", "")
    role            = request.form.get("role", "customer")
    phone           = request.form.get("phone", "").strip()
    membership_type = request.form.get("membership_type", "").strip()
    amount_paid_raw = request.form.get("amount_paid", "").strip()
    date_of_expiry  = request.form.get("date_of_expiry", "").strip()

    if role not in ("superadmin", "reseller", "customer"):
        flash("Invalid role.", "danger")
        return redirect(url_for("admin_users"))
    if not username or not password:
        flash("Username and password are required.", "danger")
        return redirect(url_for("admin_users"))

    reseller_id_raw = request.form.get("reseller_id", "").strip()
    reseller_id_val = int(reseller_id_raw) if reseller_id_raw else None
    if role != "customer":
        reseller_id_val = None

    amount_paid_val = float(amount_paid_raw) if amount_paid_raw else None

    result = database.create_user(
        username, email, password, role, reseller_id_val,
        phone=phone or None,
        membership_type=membership_type or None,
        amount_paid=amount_paid_val,
        date_of_expiry=date_of_expiry or None,
    )
    if result is None:
        flash(f"Username '{username}' already exists.", "danger")
    else:
        flash(f"User '{username}' ({role}) created successfully.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@superadmin_required
def admin_delete_user(user_id):
    if user_id == session["user_id"]:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("admin_users"))
    database.delete_user(user_id)
    flash("User deleted.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/change-role", methods=["POST"])
@superadmin_required
def admin_change_role(user_id):
    if user_id == session["user_id"]:
        flash("You cannot change your own role.", "danger")
        return redirect(url_for("admin_users"))
    role = request.form.get("role", "").strip()
    if role not in ("superadmin", "reseller", "customer"):
        flash("Invalid role.", "danger")
        return redirect(url_for("admin_users"))
    database.update_user_role(user_id, role)
    flash("Role updated successfully.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/assign-reseller", methods=["POST"])
@superadmin_required
def admin_assign_reseller(user_id):
    reseller_id_raw = request.form.get("reseller_id", "").strip()
    reseller_id_val = int(reseller_id_raw) if reseller_id_raw else None
    database.assign_customer_reseller(user_id, reseller_id_val)
    flash("Reseller assignment updated.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/change-password", methods=["POST"])
@superadmin_required
def admin_change_password(user_id):
    new_password = request.form.get("new_password", "")
    if not new_password:
        flash("Password cannot be empty.", "danger")
        return redirect(url_for("admin_users"))
    database.update_user_password(user_id, new_password)
    flash("Password updated successfully.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/update-profile", methods=["POST"])
@superadmin_required
def admin_update_user_profile(user_id):
    phone           = request.form.get("phone", "").strip()
    membership_type = request.form.get("membership_type", "").strip()
    amount_paid     = request.form.get("amount_paid", "").strip()
    date_of_expiry  = request.form.get("date_of_expiry", "").strip()
    database.update_user_profile(user_id, phone, membership_type, amount_paid, date_of_expiry)
    flash("User profile updated.", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/add-payment", methods=["POST"])
@superadmin_required
def admin_add_payment(user_id):
    membership_type = request.form.get("membership_type", "").strip()
    amount          = request.form.get("amount", "").strip()
    payment_date    = request.form.get("payment_date", "").strip()
    notes           = request.form.get("notes", "").strip()
    database.add_payment_history(user_id, membership_type, amount, payment_date, notes)
    flash("Payment record added.", "success")
    return redirect(url_for("admin_users"))


@app.route("/api/admin/users/<int:user_id>/payment-history")
@superadmin_required
def api_user_payment_history(user_id):
    return jsonify(database.get_all_payment_history(user_id))


# ── Public API ────────────────────────────────────────────────────────────────

@app.route("/api/leads", methods=["POST"])
def api_save_lead():
    data       = request.get_json() or {}
    name       = (data.get("name") or "").strip()
    email      = (data.get("email") or "").strip()
    phone      = (data.get("phone") or "").strip()
    profession = (data.get("profession") or "").strip()
    if not name or not email:
        return jsonify({"error": "Name and email are required."}), 400
    database.save_lead(name, email, phone, profession)
    return jsonify({"ok": True}), 201


# ── API routes ────────────────────────────────────────────────────────────────

@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    user_id, role = current_user()
    if "files" not in request.files:
        return jsonify({"error": "No files in request"}), 400
    if not ANTHROPIC_API_KEY:
        return jsonify({"error": "ANTHROPIC_API_KEY not configured on server"}), 500

    # Optional project assignment
    project_id_raw = request.form.get("project_id", "").strip()
    project_id_val = int(project_id_raw) if project_id_raw else None
    # Verify project ownership if provided
    if project_id_val:
        proj = database.get_project_by_id(project_id_val, user_id, role)
        if not proj:
            project_id_val = None  # ignore invalid/foreign project

    files = request.files.getlist("files")
    uploaded, skipped, resume_ids = 0, 0, []

    for f in files:
        if not f.filename or not allowed_file(f.filename):
            skipped += 1
            continue
        pdf_bytes = f.read()
        resume_id = database.insert_resume(f.filename, pdf_bytes, user_id, project_id_val)
        resume_ids.append(resume_id)
        executor.submit(extractor.process_resume, resume_id, ANTHROPIC_API_KEY, database)
        uploaded += 1

    return jsonify({"uploaded": uploaded, "skipped": skipped, "ids": resume_ids})


@app.route("/api/contacts", methods=["GET"])
@login_required
def api_contacts():
    user_id, role = current_user()
    contacts      = database.get_all_contacts(user_id, role)
    search        = request.args.get("search",        "").lower().strip()
    location      = request.args.get("location",      "").strip()
    job_title     = request.args.get("job_title",     "").strip()
    skill         = request.args.get("skill",         "").lower().strip()
    project_id_s  = request.args.get("project_id",   "").strip()
    uploader_role = request.args.get("uploader_role", "").strip()

    if search:
        def matches(c):
            return any(search in (c.get(f) or "").lower() for f in
                       ["name", "email", "company", "location", "job_title",
                        "uploaded_by_username"]) or \
                   search in " ".join(c.get("skills") or []).lower()
        contacts = [c for c in contacts if matches(c)]

    if location:
        contacts = [c for c in contacts if (c.get("location") or "").strip() == location]
    if job_title:
        contacts = [c for c in contacts if (c.get("job_title") or "").strip() == job_title]
    if skill:
        contacts = [c for c in contacts
                    if any(skill == s.lower() for s in (c.get("skills") or []))]
    if project_id_s:
        try:
            pid = int(project_id_s)
            contacts = [c for c in contacts if c.get("project_id") == pid]
        except ValueError:
            pass
    if uploader_role and role == "superadmin":
        contacts = [c for c in contacts
                    if (c.get("uploaded_by_role") or "") == uploader_role]
    customer_id_s = request.args.get("customer_id", "").strip()
    if customer_id_s and role == "reseller":
        try:
            cid = int(customer_id_s)
            contacts = [c for c in contacts if c.get("uploaded_by_user_id") == cid]
        except ValueError:
            pass

    return jsonify(contacts)


@app.route("/api/filter-options", methods=["GET"])
@login_required
def api_filter_options():
    user_id, role = current_user()
    return jsonify(database.get_filter_options(user_id, role))


@app.route("/api/status", methods=["GET"])
@login_required
def api_status():
    user_id, role = current_user()
    return jsonify(database.get_processing_status(user_id, role))


@app.route("/api/export", methods=["GET"])
@login_required
def api_export():
    user_id, role = current_user()
    contacts = database.get_all_contacts(user_id, role)
    if not contacts:
        return jsonify({"error": "No contacts to export yet"}), 404

    rows = []
    for c in contacts:
        skills_list = c.get("skills") or []
        if isinstance(skills_list, str):
            try:
                skills_list = json.loads(skills_list)
            except Exception:
                skills_list = []

        other = {}
        try:
            raw_other = c.get("other_details") or "{}"
            other = json.loads(raw_other) if isinstance(raw_other, str) else (raw_other or {})
        except Exception:
            pass

        row = {
            "Name":             c.get("name") or "",
            "Email":            c.get("email") or "",
            "Phone":            c.get("phone") or "",
            "LinkedIn":         c.get("linkedin") or "",
            "Location":         c.get("location") or "",
            "Job Title":        c.get("job_title") or "",
            "Company":          c.get("company") or "",
            "Skills":           ", ".join(skills_list),
            "GitHub":           other.get("github") or "",
            "Portfolio":        other.get("portfolio") or "",
            "Education":        other.get("education") or "",
            "Years Experience": other.get("years_experience") or "",
            "Source File":      c.get("original_filename") or "",
            "Upload Date":      str(c.get("upload_date") or ""),
            "Extracted At":     str(c.get("extracted_at") or ""),
        }
        if role == "superadmin":
            row["Uploaded By"] = c.get("uploaded_by_username") or ""
            row["User Role"]   = c.get("uploaded_by_role") or ""
        rows.append(row)

    df = pd.DataFrame(rows)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Contacts", index=False)
        ws = writer.sheets["Contacts"]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Make LinkedIn cells clickable hyperlinks
        linkedin_col = None
        for cell in ws[1]:
            if cell.value == "LinkedIn":
                linkedin_col = cell.column_letter
                break
        if linkedin_col:
            for row_cells in ws.iter_rows(
                min_row=2,
                min_col=ws[f"{linkedin_col}1"].column,
                max_col=ws[f"{linkedin_col}1"].column
            ):
                cell = row_cells[0]
                url = cell.value
                if url and url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")

        for col_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resume_contacts.xlsx",
    )


@app.route("/api/projects", methods=["GET"])
@login_required
def api_get_projects():
    user_id, role = current_user()
    return jsonify(database.get_all_projects(user_id, role))


@app.route("/api/projects", methods=["POST"])
@login_required
def api_create_project():
    user_id, role = current_user()
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    description = (data.get("description") or "").strip()
    if not name:
        return jsonify({"error": "Project name is required"}), 400
    project = database.create_project(name, description, user_id)
    if project is None:
        return jsonify({"error": "Failed to create project"}), 500
    return jsonify(project), 201


@app.route("/api/projects/<int:project_id>", methods=["DELETE"])
@login_required
def api_delete_project(project_id):
    user_id, role = current_user()
    project = database.get_project_by_id(project_id, user_id, role)
    if not project:
        return jsonify({"error": "Not found or access denied"}), 404
    database.delete_project(project_id)
    return jsonify({"deleted": project_id})


@app.route("/api/projects/<int:project_id>/contacts", methods=["GET"])
@login_required
def api_project_contacts(project_id):
    user_id, role = current_user()
    project = database.get_project_by_id(project_id, user_id, role)
    if not project:
        return jsonify({"error": "Not found"}), 404

    contacts      = database.get_project_contacts(project_id, user_id, role)
    search        = request.args.get("search",        "").lower().strip()
    location      = request.args.get("location",      "").strip()
    job_title     = request.args.get("job_title",     "").strip()
    skill         = request.args.get("skill",         "").lower().strip()
    uploader_role = request.args.get("uploader_role", "").strip()

    if search:
        def matches(c):
            return any(search in (c.get(f) or "").lower() for f in
                       ["name", "email", "company", "location", "job_title",
                        "uploaded_by_username"]) or \
                   search in " ".join(c.get("skills") or []).lower()
        contacts = [c for c in contacts if matches(c)]
    if location:
        contacts = [c for c in contacts if (c.get("location") or "").strip() == location]
    if job_title:
        contacts = [c for c in contacts if (c.get("job_title") or "").strip() == job_title]
    if skill:
        contacts = [c for c in contacts
                    if any(skill == s.lower() for s in (c.get("skills") or []))]
    if uploader_role and role == "superadmin":
        contacts = [c for c in contacts
                    if (c.get("uploaded_by_role") or "") == uploader_role]
    customer_id_s = request.args.get("customer_id", "").strip()
    if customer_id_s and role == "reseller":
        try:
            cid = int(customer_id_s)
            contacts = [c for c in contacts if c.get("uploaded_by_user_id") == cid]
        except ValueError:
            pass

    return jsonify(contacts)


@app.route("/api/projects/<int:project_id>/filter-options", methods=["GET"])
@login_required
def api_project_filter_options(project_id):
    user_id, role = current_user()
    project = database.get_project_by_id(project_id, user_id, role)
    if not project:
        return jsonify({"error": "Not found"}), 404
    contacts = database.get_project_contacts(project_id, user_id, role)
    locations  = sorted({(c.get("location") or "").strip() for c in contacts if c.get("location")})
    job_titles = sorted({(c.get("job_title") or "").strip() for c in contacts if c.get("job_title")})
    skills_set = set()
    for c in contacts:
        for s in (c.get("skills") or []):
            if s:
                skills_set.add(s.strip())
    uploader_roles = []
    if role == "superadmin":
        uploader_roles = sorted({c.get("uploaded_by_role") for c in contacts
                                   if c.get("uploaded_by_role")})
    return jsonify({
        "locations":      locations,
        "job_titles":     job_titles,
        "skills":         sorted(skills_set),
        "uploader_roles": uploader_roles,
    })


@app.route("/api/projects/<int:project_id>/export", methods=["GET"])
@login_required
def api_project_export(project_id):
    user_id, role = current_user()
    project = database.get_project_by_id(project_id, user_id, role)
    if not project:
        return jsonify({"error": "Not found"}), 404

    contacts = database.get_project_contacts(project_id, user_id, role)
    if not contacts:
        return jsonify({"error": "No contacts in this project yet"}), 404

    rows = []
    for c in contacts:
        skills_list = c.get("skills") or []
        if isinstance(skills_list, str):
            try:
                skills_list = json.loads(skills_list)
            except Exception:
                skills_list = []

        other = {}
        try:
            raw_other = c.get("other_details") or "{}"
            other = json.loads(raw_other) if isinstance(raw_other, str) else (raw_other or {})
        except Exception:
            pass

        row = {
            "Name":             c.get("name") or "",
            "Email":            c.get("email") or "",
            "Phone":            c.get("phone") or "",
            "LinkedIn":         c.get("linkedin") or "",
            "Location":         c.get("location") or "",
            "Job Title":        c.get("job_title") or "",
            "Company":          c.get("company") or "",
            "Skills":           ", ".join(skills_list),
            "GitHub":           other.get("github") or "",
            "Portfolio":        other.get("portfolio") or "",
            "Education":        other.get("education") or "",
            "Years Experience": other.get("years_experience") or "",
            "Source File":      c.get("original_filename") or "",
            "Upload Date":      str(c.get("upload_date") or ""),
            "Extracted At":     str(c.get("extracted_at") or ""),
        }
        if role == "superadmin":
            row["Uploaded By"] = c.get("uploaded_by_username") or ""
            row["User Role"]   = c.get("uploaded_by_role") or ""
        rows.append(row)

    df = pd.DataFrame(rows)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Contacts", index=False)
        ws = writer.sheets["Contacts"]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        linkedin_col = None
        for cell in ws[1]:
            if cell.value == "LinkedIn":
                linkedin_col = cell.column_letter
                break
        if linkedin_col:
            for row_cells in ws.iter_rows(
                min_row=2,
                min_col=ws[f"{linkedin_col}1"].column,
                max_col=ws[f"{linkedin_col}1"].column
            ):
                cell = row_cells[0]
                url = cell.value
                if url and url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")

        for col_cells in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

    output.seek(0)
    safe_name = "".join(c for c in project["name"] if c.isalnum() or c in " _-").strip()
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{safe_name}_contacts.xlsx",
    )


@app.route("/api/resumes/failed", methods=["GET"])
@login_required
def api_failed():
    user_id, role = current_user()
    return jsonify(database.get_failed_resumes(user_id, role))


@app.route("/api/resumes/<int:resume_id>", methods=["DELETE"])
@login_required
def api_delete_resume(resume_id):
    user_id, role = current_user()
    resume = database.get_resume_by_id(resume_id, user_id, role)
    if not resume:
        return jsonify({"error": "Not found"}), 404
    database.delete_resume(resume_id)
    return jsonify({"deleted": resume_id})


@app.route("/uploads/<int:resume_id>")
@login_required
def serve_pdf(resume_id):
    user_id, role = current_user()
    resume = database.get_resume_by_id(resume_id, user_id, role)
    if not resume:
        return "Not found", 404
    pdf_bytes = database.get_pdf_bytes(resume_id)
    if not pdf_bytes:
        return "PDF not found", 404
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=False,
        download_name=resume.get("original_filename", "resume.pdf"),
    )


# ── Startup ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if not ANTHROPIC_API_KEY:
        print("\nWARNING: ANTHROPIC_API_KEY not set.\n")
    print("Starting Resume Extractor at http://localhost:5000")
    app.run(debug=True, host="0.0.0.0", port=5000, use_reloader=False)
